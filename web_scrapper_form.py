# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('top_news.xlsx')

# create the sheet object
sheet = wb.active


def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 50
    sheet.column_dimensions['E'].width = 50

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Website"
    sheet.cell(row=1, column=2).value = "Date"
    sheet.cell(row=1, column=3).value = "Author"
    sheet.cell(row=1, column=4).value = "News Description"
    sheet.cell(row=1, column=5).value = "News Link"


# Function to set focus (cursor)
def focus1(event):
    # set focus on the course_field box
    username_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    password_field.focus_set()


# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    username_field.delete(0, END)
    password_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (username_field.get() != "admin" or password_field.get() != "admin"):
        print("Username and or Password not correct")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column
        import requests
        from bs4 import BeautifulSoup

        websites = {'URLs': ['https://tribune.com.pk/', 'https://www.dawn.com'],
                    'attrs1': ['story', 'border-3'],
                    'attrs2': ['story clearfix', 'template__header'],
                    'desc': ['h1', 'h2']}

        records = []

        for i in range(len(websites['URLs'])):
            r = requests.get(websites['URLs'][i])
            soup = BeautifulSoup(r.text, 'html.parser')
            results = soup.find_all('div', attrs={'class': websites['attrs1'][i]})

            if i == 1:
                results = results[0].find_all('article')
            url = str(results[0].find('a')['href'])

            r = requests.get(url)
            soup = BeautifulSoup(r.text, 'html.parser')
            results = soup.find_all('div', attrs={'class': websites['attrs2'][i]})

            website = websites['URLs'][i]
            link = url
            description = results[0].find(websites['desc'][i]).text
            if i == 0:
                date = results[0].find_all('div')[3].text.split()
                author = results[0].find_all('div')[2].text.split()
                author.pop(0)
                author = ' '.join(author)
            elif i == 1:
                date = results[0].find('div')('span')[2].text.split(' ')
                author = results[0].find('div')('span')[0].text
            date.pop(0)
            date = ' '.join(date)

            records.append(list((website, date, author, description, link)))

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        for i in range(len(records)):
            sheet.cell(row=current_row + 1+i, column=1).value = records[i][0]
            sheet.cell(row=current_row + 1+i, column=2).value = records[i][1]
            sheet.cell(row=current_row + 1+i, column=3).value = records[i][2]
            sheet.cell(row=current_row + 1+i, column=4).value = records[i][3]
            sheet.cell(row=current_row + 1+i, column=5).value = records[i][4]
        # save the file
        wb.save('top_news.xlsx')

        # call the clear() function
        clear()
        root.destroy()


if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("Top News Form")

    # set the configuration of GUI window
    root.geometry("500x300")

    excel()

    # create a Form label
    heading = Label(root, text="Form", bg="light green")

    # create a Name label
    username = Label(root, text="Username", bg="light green")

    # create a Course label
    password = Label(root, text="Password", bg="light green")


    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    username.grid(row=1, column=0)
    password.grid(row=2, column=0)

    # create a text entry box
    # for typing the information
    username_field = Entry(root)
    password_field = Entry(root)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    username_field.grid(row=1, column=1, ipadx="100")
    password_field.grid(row=2, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=8, column=1)

    # start the GUI
    root.mainloop()
